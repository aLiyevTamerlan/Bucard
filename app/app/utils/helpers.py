import json
import qrcode
from io import BytesIO
from openpyxl import load_workbook
from django.core.files import File

from core.models import Category

def generate_qr_code_and_save(obj):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
        )
    qr.add_data(obj.get_full_path())
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    blob = BytesIO()
    img.save(blob, 'PNG')
    obj.qr_code.save(f'{obj.name}.png', File(blob), save=False)
    obj.save()

    return obj

def export_categories_from_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=10, max_col=200, values_only=True):
        if row[4]:
            name = row[4].replace('"', '').capitalize().replace(',', '')
            Category.objects.get_or_create(name=name)
    return True


def export_categories_from_json(file_path):
    """
    Export categories from json file
    """

    with open(file_path) as f:
        data = json.load(f)
        for item in data:
            category = Category.objects.create(name=item)

    return True